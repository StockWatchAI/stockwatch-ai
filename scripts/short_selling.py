#!/usr/bin/env python3
"""JPXが毎営業日公表する空売り残高を取り込み、増えた銘柄を通知する。

    python3 scripts/short_selling.py --inspect   # 取得だけして構造を出す（認証不要）
    python3 scripts/short_selling.py             # 取り込んで通知まで送る

日本にしかない材料である。米国には個別ファンドの売り建てを実名で公表させる
制度が無く、残高割合0.5%以上を毎営業日17時に、法人は商号つきで出しているのは
JPXだけ。株価やニュースと違って他のアプリが持っていないため、ここが効く。

**Excelの形を決め打ちにしない。** JPXの様式は列の増減や見出しの表記ゆれがある。
列番号で拾うと、様式が変わった日から黙って別の列を読み続けることになる。
見出しの文字で列を探し、必要な列が見つからなければ**何も書かずに終了する**。
market_close.py と同じで、誤った値を送るより送らないほうがよい。

前回の残高はFirestoreに置いた前回ぶんと比べる。比較対象が無い初回は
順位だけを保存し、通知は送らない（全銘柄が「新規」になってしまうため）。
"""

import argparse
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone

import requests

INDEX_URL = "https://www.jpx.co.jp/markets/public/short-selling/index.html"
ORIGIN = "https://www.jpx.co.jp"

# Firestoreの置き場所。usersと違い全員で共有する1件
SNAPSHOT_DOC = ("market", "short_selling")

# 通知に載せる上限。多いと通知の本文に収まらない
TOP_N = 5

# 「増えた」と見なす残高割合の増分（ポイント）。
# 0.5%以上で報告義務が生じるので、0.3ポイントは1社が新たに報告し始めた程度の変化にあたる
INCREASE_THRESHOLD = 0.3

JST = timezone(timedelta(hours=9))

# 見出しの表記ゆれを吸収する。左から順に試し、最初に当たった列を使う。
# 「空売り残高割合」より先に「直近」を判定しないよう、除外語も持つ
COLUMNS = {
    "code": {"any": ["銘柄コード", "コード"], "not": []},
    "name": {"any": ["銘柄名", "銘柄"], "not": ["コード"]},
    "holder": {"any": ["商号", "名称", "氏名", "空売り者"], "not": []},
    "ratio": {"any": ["空売り残高割合", "残高割合"], "not": ["直近", "前回"]},
    "quantity": {"any": ["空売り残高数量", "残高数量"], "not": ["直近", "前回"]},
    "date": {"any": ["計算年月日"], "not": ["直近", "前回"]},
}
REQUIRED = ["code", "ratio"]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 取得
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def get(url, **kwargs):
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=60, **kwargs)
    r.raise_for_status()
    return r


def find_report_links(html):
    """一覧ページからExcelへのリンクを拾う。新しいものが上にある前提は置かない。"""
    links = []
    for match in re.finditer(r'href="([^"]+\.xlsx?)"', html, re.IGNORECASE):
        href = match.group(1)
        url = href if href.startswith("http") else ORIGIN + href
        # ファイル名に入っている8桁の日付を順序の手がかりにする
        stamp = re.search(r"(20\d{6})", href)
        links.append({"url": url, "date": stamp.group(1) if stamp else ""})
    # 同じファイルが複数箇所から貼られていることがある
    seen, unique = set(), []
    for link in links:
        if link["url"] in seen:
            continue
        seen.add(link["url"])
        unique.append(link)
    return unique


def pick_latest(links):
    """日付つきのうち最も新しいものを選ぶ。日付が無いものは残高一覧ではない可能性が高い。"""
    dated = [l for l in links if l["date"]]
    if not dated:
        return None
    return max(dated, key=lambda l: l["date"])


def load_rows(content, url):
    """Excelを行の配列にする。xlsx（openpyxl）と旧形式のxls（xlrd）の両方を受ける。"""
    if url.lower().endswith(".xlsx"):
        import openpyxl

        book = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = book[book.sheetnames[0]]
        return book.sheetnames, [list(r) for r in sheet.iter_rows(values_only=True)]

    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    rows = [sheet.row_values(i) for i in range(sheet.nrows)]
    return book.sheet_names(), rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 解釈
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def normalize(value):
    """見出しの比較用。改行・空白・全角空白を落とす。"""
    if value is None:
        return ""
    return re.sub(r"[\s　]+", "", str(value))


def find_header(rows, limit=20):
    """見出しの行と、列の対応を探す。

    先頭に表題や注記が入るため、行番号は決め打ちにできない。
    必要な列が最も多く見つかった行を見出しとみなす。
    """
    best = None
    for index, row in enumerate(rows[:limit]):
        mapping = {}
        for key, rule in COLUMNS.items():
            for column, cell in enumerate(row):
                text = normalize(cell)
                if not text:
                    continue
                if any(ng in text for ng in rule["not"]):
                    continue
                if any(ok in text for ok in rule["any"]):
                    mapping.setdefault(key, column)
                    break
        if best is None or len(mapping) > len(best[1]):
            best = (index, mapping)
    return best if best else (None, {})


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = re.sub(r"[,\s　%％]", "", str(value))
    try:
        return float(text)
    except ValueError:
        return None


def to_code(value):
    """銘柄コードを4桁の文字列にする。Excelは数値で持っていることが多い。"""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = re.sub(r"[\s　]", "", str(value))
    # 5桁の新コードや英字を含むコードもあるため、桁数では弾かない
    return text if re.fullmatch(r"[0-9A-Z]{4,5}", text) else None


def parse(rows, header_index, mapping):
    """見出しの次の行から、1報告＝1件として読む。"""
    records = []
    for row in rows[header_index + 1:]:
        code = to_code(row[mapping["code"]]) if mapping["code"] < len(row) else None
        if not code:
            continue
        ratio = to_number(row[mapping["ratio"]]) if mapping["ratio"] < len(row) else None
        if ratio is None:
            continue
        # 割合は0.0052のような小数で入っていることも、0.52のように％で
        # 入っていることもある。報告義務が0.5%以上なので、1未満が並ぶ列は小数とみなす
        records.append(
            {
                "code": code,
                "name": _cell(row, mapping, "name"),
                "holder": _cell(row, mapping, "holder"),
                "ratio": ratio,
                "quantity": to_number(row[mapping["quantity"]])
                if "quantity" in mapping and mapping["quantity"] < len(row)
                else None,
            }
        )
    return records


def _cell(row, mapping, key):
    if key not in mapping or mapping[key] >= len(row):
        return ""
    value = row[mapping[key]]
    return "" if value is None else str(value).strip()


def rescale_ratios(records):
    """割合の単位をパーセントに揃える。

    報告義務は残高割合0.5%以上なので、パーセント表記なら0.5未満はほぼ現れない。
    値がすべて1未満なら小数表記だと判断して100倍する。
    """
    ratios = [r["ratio"] for r in records if r["ratio"] is not None]
    if ratios and max(ratios) < 1.0:
        for r in records:
            r["ratio"] *= 100
    return records


def aggregate(records):
    """銘柄ごとにまとめる。残高割合は報告者ごとに出るので合計が全体の売り建て比率になる。"""
    by_code = {}
    for r in records:
        entry = by_code.setdefault(
            r["code"], {"code": r["code"], "name": r["name"], "ratio": 0.0, "holders": []}
        )
        entry["ratio"] += r["ratio"]
        if r["name"] and not entry["name"]:
            entry["name"] = r["name"]
        if r["holder"]:
            entry["holders"].append({"name": r["holder"], "ratio": round(r["ratio"], 2)})
    for entry in by_code.values():
        entry["ratio"] = round(entry["ratio"], 2)
        # 大きく持っている順。通知に出すのは先頭だけ
        entry["holders"].sort(key=lambda h: h["ratio"], reverse=True)
        entry["holders"] = entry["holders"][:5]
    return by_code


def rank_increases(current, previous):
    """前回からの増加が大きい順に並べる。前回が無い銘柄は新規として増分＝全量とみなす。"""
    rows = []
    for code, entry in current.items():
        before = previous.get(code, {}).get("ratio", 0.0)
        delta = entry["ratio"] - before
        if delta < INCREASE_THRESHOLD:
            continue
        rows.append({**entry, "previous": round(before, 2), "delta": round(delta, 2)})
    rows.sort(key=lambda r: r["delta"], reverse=True)
    return rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 確認用（認証なしで動かせる）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def inspect():
    """取り込みの前段だけを動かして、実物の構造を出す。

    JPXの様式は手元から確認できないため、まずこれをCIで走らせて
    列の見出しと値の入り方を確かめる。
    """
    print(f"一覧ページ: {INDEX_URL}")
    html = get(INDEX_URL).text
    print(f"HTML長: {len(html)}")

    links = find_report_links(html)
    print(f"\nExcelへのリンク {len(links)} 件（先頭10件）:")
    for link in links[:10]:
        print(f"  date={link['date'] or '(なし)':10} {link['url']}")

    latest = pick_latest(links)
    if not latest:
        print("\n日付つきのExcelが見つからない。リンクの抽出条件を見直す必要がある")
        return 1
    print(f"\n選んだファイル: {latest['url']}")

    content = get(latest["url"]).content
    print(f"サイズ: {len(content)} バイト")

    sheets, rows = load_rows(content, latest["url"])
    print(f"シート: {sheets}")
    print(f"行数: {len(rows)}")

    print("\n--- 先頭12行（各行の先頭12列） ---")
    for i, row in enumerate(rows[:12]):
        cells = [normalize(c)[:18] for c in row[:12]]
        print(f"[{i:2}] {cells}")

    header_index, mapping = find_header(rows)
    print(f"\n見出しとみなした行: {header_index}")
    print(f"列の対応: {mapping}")
    missing = [k for k in REQUIRED if k not in mapping]
    if missing:
        print(f"必須の列が見つからない: {missing} → COLUMNS の語を追加する")
        return 1

    records = rescale_ratios(parse(rows, header_index, mapping))
    print(f"\n読めた報告件数: {len(records)}")
    for r in records[:5]:
        print(f"  {r}")

    current = aggregate(records)
    print(f"\n銘柄数: {len(current)}")
    top = sorted(current.values(), key=lambda e: e["ratio"], reverse=True)[:10]
    print("残高割合が大きい銘柄:")
    for e in top:
        holders = "／".join(h["name"][:14] for h in e["holders"][:2])
        print(f"  {e['code']} {e['name'][:16]:16} {e['ratio']:5.2f}%  {holders}")
    return 0


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 保存と通知
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def load_previous(db):
    doc = db.collection(SNAPSHOT_DOC[0]).document(SNAPSHOT_DOC[1]).get()
    if not doc.exists:
        return None, None
    data = doc.to_dict() or {}
    return data.get("stocks") or {}, data.get("sourceDate")


def save_snapshot(db, stocks, source_url, source_date, ranking):
    db.collection(SNAPSHOT_DOC[0]).document(SNAPSHOT_DOC[1]).set(
        {
            "stocks": stocks,
            "ranking": ranking[:50],
            "sourceUrl": source_url,
            "sourceDate": source_date,
            "updatedAt": datetime.now(JST).isoformat(),
        }
    )


def notify(db, ranking):
    """自分が登録している銘柄が増えていた人にだけ送る。

    全員に同じ順位表を送ると、持っていない銘柄の話ばかりになる。
    watchlistは `7203.T` 形式、JPXは `7203` なので、末尾を落として突き合わせる。
    """
    by_code = {r["code"]: r for r in ranking}
    users = list(db.collection("users").stream())
    print(f"ユーザー数: {len(users)}")

    sent = failed = skipped = 0
    for u in users:
        data = u.to_dict()
        token = data.get("fcmToken")
        if not token:
            continue
        if data.get("notifyShortSelling", True) is False:
            skipped += 1
            continue

        hits = []
        for symbol in data.get("watchlist") or []:
            if not symbol.upper().endswith(".T"):
                continue
            hit = by_code.get(symbol[:-2])
            if hit:
                hits.append(hit)
        if not hits:
            continue

        hits.sort(key=lambda r: r["delta"], reverse=True)
        hits = hits[:TOP_N]
        title = "空売り残高が増えた銘柄があります"
        body = "\n".join(
            f"{h['name'] or h['code']} {h['previous']:.2f}% → {h['ratio']:.2f}%（+{h['delta']:.2f}）"
            for h in hits
        )
        try:
            res = messaging.send(
                messaging.Message(
                    notification=messaging.Notification(title=title, body=body),
                    data={"symbol": f"{hits[0]['code']}.T"},
                    token=token,
                )
            )
            sent += 1
            print(f"通知送信成功: {u.id} → FCM応答={res}")
        except Exception as e:
            failed += 1
            print(f"通知送信失敗: {u.id} → {repr(e)}")

    print(f"送信 {sent} 件 / 失敗 {failed} 件 / 設定オフ {skipped} 件")


def main():
    html = get(INDEX_URL).text
    latest = pick_latest(find_report_links(html))
    if not latest:
        print("Excelのリンクが見つからないため中止")
        return

    content = get(latest["url"]).content
    _, rows = load_rows(content, latest["url"])
    header_index, mapping = find_header(rows)
    missing = [k for k in REQUIRED if k not in mapping]
    if header_index is None or missing:
        # 様式が変わったときに、別の列を読んで見当違いの数字を送るのを防ぐ
        print(f"必要な列が見つからないため中止: {missing}")
        return

    records = rescale_ratios(parse(rows, header_index, mapping))
    if not records:
        print("1件も読めなかったため中止")
        return
    current = aggregate(records)
    print(f"{latest['date']}: 報告 {len(records)} 件 / 銘柄 {len(current)} 件")

    previous, previous_date = load_previous(db)

    if previous is None:
        print("前回ぶんが無いため、保存だけして通知は送らない")
        save_snapshot(db, current, latest["url"], latest["date"], [])
        return

    if previous_date == latest["date"]:
        print(f"公表日 {latest['date']} は取り込み済みのため中止（休場か更新前）")
        return

    ranking = rank_increases(current, previous)
    print(f"増加した銘柄: {len(ranking)} 件")
    for r in ranking[:10]:
        print(f"  {r['code']} {r['name'][:16]:16} {r['previous']:.2f}% → {r['ratio']:.2f}%")

    save_snapshot(db, current, latest["url"], latest["date"], ranking)

    if os.environ.get("DRY_RUN") == "1":
        print("DRY_RUN のため通知は送りません")
        return
    notify(db, ranking)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--inspect", action="store_true", help="取得と解釈だけ試す（認証不要）")
    args = parser.parse_args()

    if args.inspect:
        sys.exit(inspect())

    import firebase_admin
    from firebase_admin import credentials, firestore, messaging

    cred = credentials.Certificate(json.loads(os.environ["FIREBASE_SERVICE_ACCOUNT"]))
    firebase_admin.initialize_app(cred)
    db = firestore.client()
    main()
